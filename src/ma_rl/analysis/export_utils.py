from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def write_excel_friendly_csv(path: str | Path, rows: list[dict]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if not rows:
        return

    with path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=list(rows[0].keys()),
            delimiter=";",
        )
        writer.writeheader()
        writer.writerows(rows)


def write_simple_xlsx(path: str | Path, rows: list[dict], sheet_name: str = "Results") -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)

    if not rows:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(rows[0].keys())
    ws.append(headers)

    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in rows:
        ws.append([row.get(header) for header in headers])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row_idx in range(2, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=column_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(column_idx)].width = min(max_len + 2, 40)

    wb.save(path)